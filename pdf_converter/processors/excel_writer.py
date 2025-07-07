"""Excel writer for converting PDF data to Excel files."""

from __future__ import annotations

import time
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    Workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None

from ..models import ProcessingResult
from ..utils import generate_output_filename, sanitize_filename
from .base import BaseProcessor
from .structured_excel_writer import StructuredExcelWriter


class ExcelWriter(BaseProcessor):
    """Excel writer for converting PDF data to Excel files."""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None) -> None:
        """Initialize the Excel writer.
        
        Args:
            config: Configuration dictionary
        """
        super().__init__(config)
        self._check_dependencies()
    
    def _check_dependencies(self) -> None:
        """Check if required dependencies are available."""
        if pd is None:
            raise ImportError("pandas is required for Excel writing")
        if Workbook is None:
            raise ImportError("openpyxl is required for Excel writing")
    
    def can_process(self, file_path: str) -> bool:
        """Check if this processor can handle the given file."""
        # This processor doesn't process files directly, it writes Excel files
        return False
    
    def process(self, job: Any) -> Any:
        """This method is not used for Excel writer."""
        raise NotImplementedError("ExcelWriter does not process files directly")
    
    def write_to_excel(self, result: ProcessingResult, output_path: Optional[Path] = None) -> Path:
        """Write processing result to Excel file.
        
        Args:
            result: Processing result containing extracted data
            output_path: Optional output path, will generate if not provided
            
        Returns:
            Path to the created Excel file
        """
        if output_path is None:
            output_path = self._generate_output_path(result)
        
        # Use StructuredExcelWriter if report_type is 'DD'
        report_type = getattr(result.job.excel_options, 'report_type', 'Standard')
        if report_type == 'DD':
            return StructuredExcelWriter().write_structured_excel(result, output_path)

        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Write data based on content type
        if result.extracted_data.get("tables"):
            self._write_tables(wb, result)
        
        if result.extracted_data.get("text"):
            self._write_text(wb, result)
        
        # Write metadata if requested
        if result.job.excel_options.include_metadata:
            self._write_metadata(wb, result)
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
    
    def _generate_output_path(self, result: ProcessingResult) -> Path:
        """Generate output file path."""
        output_dir = Path.cwd()  # Default to current working directory
        output_dir.mkdir(parents=True, exist_ok=True)
        
        return generate_output_filename(
            result.job.input_file,
            result.job.excel_options.format,
            output_dir
        )
    
    def _write_tables(self, wb: Workbook, result: ProcessingResult) -> None:
        """Write tables to Excel workbook."""
        tables = result.extracted_data.get("tables", {}).get("tables", [])
        
        if not tables:
            return
        
        if result.job.excel_options.separate_sheets:
            # Create separate sheet for each table
            for i, table_info in enumerate(tables):
                sheet_name = self._generate_sheet_name(f"Table_{i+1}", wb)
                ws = wb.create_sheet(title=sheet_name)
                self._write_table_to_sheet(ws, table_info, result)
        else:
            # Write all tables to a single sheet
            sheet_name = self._generate_sheet_name("Tables", wb)
            ws = wb.create_sheet(title=sheet_name)
            
            current_row = 1
            for i, table_info in enumerate(tables):
                if i > 0:
                    current_row += 2  # Add spacing between tables
                
                # Add table header
                ws.cell(row=current_row, column=1, value=f"Table {i+1} (Page {table_info['page']})")
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
                current_row += 1
                
                # Write table data
                current_row = self._write_table_data(ws, table_info["data"], current_row, result)
    
    def _write_table_to_sheet(self, ws: Any, table_info: Dict[str, Any], result: ProcessingResult) -> None:
        """Write a single table to a worksheet."""
        # Add table header
        ws.cell(row=1, column=1, value=f"Table from Page {table_info['page']}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Write table data
        self._write_table_data(ws, table_info["data"], 3, result)
    
    def _write_table_data(self, ws: Any, table_data: List[List[str]], start_row: int, result: ProcessingResult) -> int:
        """Write table data to worksheet and return next row."""
        if not table_data:
            return start_row
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(table_data[1:], columns=table_data[0])
        
        # Write headers
        for col_idx, header in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
        
        # Auto-adjust column widths if requested
        if result.job.excel_options.auto_adjust_columns:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        
        return start_row + len(df) + 1
    
    def _write_text(self, wb: Workbook, result: ProcessingResult) -> None:
        """Write text content to Excel workbook."""
        text_data = result.extracted_data.get("text", {})
        
        if not text_data:
            return
        
        if result.job.excel_options.separate_sheets:
            # Create separate sheet for each page
            for page_num, text in text_data.get("page_texts", {}).items():
                sheet_name = self._generate_sheet_name(f"Text_Page_{page_num}", wb)
                ws = wb.create_sheet(title=sheet_name)
                self._write_text_to_sheet(ws, text, page_num, result)
        else:
            # Write all text to a single sheet
            sheet_name = self._generate_sheet_name("Text", wb)
            ws = wb.create_sheet(title=sheet_name)
            
            current_row = 1
            for page_num, text in text_data.get("page_texts", {}).items():
                if current_row > 1:
                    current_row += 2  # Add spacing between pages
                
                # Add page header
                ws.cell(row=current_row, column=1, value=f"Page {page_num}")
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                current_row += 1
                
                # Write text content
                ws.cell(row=current_row, column=1, value=text)
                current_row += 2
    
    def _write_text_to_sheet(self, ws: Any, text: str, page_num: int, result: ProcessingResult) -> None:
        """Write text content to a worksheet."""
        # Add page header
        ws.cell(row=1, column=1, value=f"Page {page_num}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        
        # Write text content
        ws.cell(row=3, column=1, value=text)
        
        # Auto-adjust column width
        if result.job.excel_options.auto_adjust_columns:
            ws.column_dimensions['A'].width = 100
    
    def _write_metadata(self, wb: Workbook, result: ProcessingResult) -> None:
        """Write metadata to Excel workbook."""
        sheet_name = self._generate_sheet_name("Metadata", wb)
        ws = wb.create_sheet(title=sheet_name)
        
        # Write metadata
        row = 1
        ws.cell(row=row, column=1, value="Processing Metadata")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        metadata = result.metadata
        for key, value in metadata.items():
            ws.cell(row=row, column=1, value=str(key))
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
    
    def _generate_sheet_name(self, base_name: str, wb: Workbook) -> str:
        """Generate a valid Excel sheet name."""
        # Excel sheet names have restrictions
        invalid_chars = r'[\\/*?:\[\]]'
        import re
        clean_name = re.sub(invalid_chars, '_', base_name)
        
        # Limit length
        if len(clean_name) > 31:
            clean_name = clean_name[:31]
        
        # Ensure uniqueness
        original_name = clean_name
        counter = 1
        while clean_name in [ws.title for ws in wb.worksheets]:
            suffix = f"_{counter}"
            clean_name = original_name[:31-len(suffix)] + suffix
            counter += 1
        
        return clean_name 