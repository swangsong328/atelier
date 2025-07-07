"""Structured Excel writer for invoice data in standardized format."""

from __future__ import annotations

import time
from pathlib import Path
from typing import Any, Dict, List, Optional
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    Workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None
    dataframe_to_rows = None

from ..models import ProcessingResult
from ..utils import generate_output_filename
from .base import BaseProcessor
from .invoice_processor import InvoiceProcessor


class StructuredExcelWriter(BaseProcessor):
    """Excel writer that outputs data in structured invoice format."""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None) -> None:
        """Initialize the structured Excel writer."""
        super().__init__(config)
        self._check_dependencies()
        self.invoice_processor = InvoiceProcessor()
    
    def _check_dependencies(self) -> None:
        """Check if required dependencies are available."""
        if pd is None:
            raise ImportError("pandas is required for structured Excel writing")
        if Workbook is None:
            raise ImportError("openpyxl is required for structured Excel writing")
    
    def can_process(self, file_path: str) -> bool:
        """Check if this processor can handle the given file."""
        # This processor doesn't process files directly, it writes Excel files
        return False
    
    def process(self, job: Any) -> Any:
        """This method is not used for Excel writer."""
        raise NotImplementedError("StructuredExcelWriter does not process files directly")
    
    def write_structured_excel(self, result: ProcessingResult, output_path: Optional[Path] = None) -> Path:
        """Write processing result to Excel file in structured format.
        
        Args:
            result: Processing result containing extracted data
            output_path: Optional output path, will generate if not provided
            
        Returns:
            Path to the created Excel file
        """
        if output_path is None:
            output_path = self._generate_output_path(result)
        
        # Transform data to structured format
        structured_df = self.invoice_processor.transform_to_structured_format(
            result.extracted_data, 
            result.metadata
        )
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create main data sheet
        ws = wb.create_sheet(title="Invoice_Data")
        
        # Write structured data
        self._write_structured_data(ws, structured_df)
        
        # Add metadata sheet if requested
        if result.job.excel_options.include_metadata:
            self._write_metadata_sheet(wb, result)
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
    
    def _generate_output_path(self, result: ProcessingResult) -> Path:
        """Generate output file path."""
        output_dir = Path.cwd()  # Default to current working directory
        output_dir.mkdir(parents=True, exist_ok=True)
        
        return generate_output_filename(
            result.job.input_file,
            result.job.excel_options.format,
            output_dir
        )
    
    def _write_structured_data(self, ws: Any, df: pd.DataFrame) -> None:
        """Write structured data to worksheet."""
        if Font is None or PatternFill is None or dataframe_to_rows is None:
            raise ImportError("openpyxl is required for structured Excel writing")
            
        if df.empty:
            # Write empty structure with headers
            headers = self.invoice_processor.required_columns
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
        else:
            # Write data with headers
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Style headers
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        # Style data rows
                        if r_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _write_metadata_sheet(self, wb: Any, result: ProcessingResult) -> None:
        """Write metadata to separate sheet."""
        if Font is None:
            raise ImportError("openpyxl is required for structured Excel writing")
            
        ws = wb.create_sheet(title="Metadata")
        
        # Write metadata
        row = 1
        ws.cell(row=row, column=1, value="Processing Metadata")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        metadata = result.metadata
        for key, value in metadata.items():
            ws.cell(row=row, column=1, value=str(key))
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50 